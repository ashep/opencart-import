#!/usr/bin/env python

import sys, re, mysql.connector, configparser
from typing import Optional, Tuple, List, Dict
from datetime import datetime
from mysql.connector.cursor import MySQLCursor
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

__author__ = 'Alexander Shepetko'
__email__ = 'a@shepetko.com'
__license__ = 'MIT'

config = configparser.ConfigParser()
config.read('config.ini')

DEBUG = config.getboolean('general', 'debug', fallback=True)

DB_NAME = config.get('db', 'name', fallback='test')
DB_HOST = config.get('db', 'host', fallback='localhost')
DB_PORT = config.getint('db', 'port', fallback=3306)
DB_USER = config.get('db', 'user', fallback='test')
DB_PASS = config.get('db', 'password', fallback='test')

EXCEL_DATA_NAMES_ROW = config.getint('excel', 'data_names_row', fallback=2)
EXCEL_DATA_START_ROW = config.getint('excel', 'data_start_row', fallback=3)

OC_LANGUAGE_ID = config.getint('opencart', 'language_id', fallback=2)
OC_STORE_ID = config.getint('opencart', 'store_id', fallback=0)
OC_LAYOUT_ID = config.getint('opencart', 'layout_id', fallback=0)
OC_STOCK_STATUS_ID = config.getint('opencart', 'stock_status_id', fallback=7)
OC_PRODUCT_QUANTITY = config.getint('opencart', 'product_quantity', fallback=10)
OC_PRODUCT_MINIMUM = config.getint('opencart', 'product_minimum', fallback=1)
OC_PRODUCT_SHIPPING = config.getint('opencart', 'product_shipping', fallback=1)
OC_PRODUCT_SUBTRACT = config.getint('opencart', 'product_subtract', fallback=0)
OC_WEIGHT_CLASS_ID = config.getint('opencart', 'weight_class_id', fallback=1)
OC_LENGTH_CLASS_ID = config.getint('opencart', 'length_class_id', fallback=2)

many_spaces_re = re.compile('\s{2,}')
db_connection = None  # type: mysql.connector.MySQLConnection


def debug(msg: str):
    if DEBUG:
        print(msg)


def transliterate(text: str) -> str:
    """Transliterate a string.
    """
    cyrillic = [
        "Щ", "щ", 'Ё', 'Ж', 'Х', 'Ц', 'Ч', 'Ш', 'Ю', 'Я',
        'ё', 'ж', 'х', 'ц', 'ч', 'ш', 'ю', 'я', 'А', 'Б',
        'В', 'Г', 'Д', 'Е', 'З', 'И', 'Й', 'К', 'Л', 'М',
        'Н', 'О', 'П', 'Р', 'С', 'Т', 'У', 'Ф', 'Ь', 'Ы',
        'Ъ', 'Э', 'а', 'б', 'в', 'г', 'д', 'е', 'з', 'и',
        'і', 'й', 'к', 'л', 'м', 'н', 'о', 'п', 'р', 'с',
        'т', 'у', 'ф', 'ь', 'ы', 'ъ', 'э', 'Ї', 'ї', 'Є',
        'є', 'Ґ', 'ґ']

    roman = [
        "Sch", "sch", 'Yo', 'Zh', 'Kh', 'Ts', 'Ch', 'Sh', 'Yu', 'Ya',
        'yo', 'zh', 'kh', 'ts', 'ch', 'sh', 'yu', 'ya', 'A', 'B',
        'V', 'G', 'D', 'E', 'Z', 'I', 'Y', 'K', 'L', 'M',
        'N', 'O', 'P', 'R', 'S', 'T', 'U', 'F', '', 'Y',
        '', 'E', 'a', 'b', 'v', 'g', 'd', 'e', 'z', 'i',
        'i', 'y', 'k', 'l', 'm', 'n', 'o', 'p', 'r', 's',
        't', 'u', 'f', '', 'y', '', 'e', 'i', 'i', 'Ye',
        'ye', 'G', 'g'
    ]

    r = ''
    for ch in text:
        try:
            i = cyrillic.index(ch)
            r += roman[i]
        except ValueError:
            r += ch

    return r


def transform_str(s: str) -> str:
    mapping = {
        '!': '', '@': '', '#': '', '$': '', '%': '', '^': '', '&': '', '*': '', '(': '', ')': '', '_': '',
        '=': '', '+': '', '"': '', "'": '', '{': '', '}': '', '[': '', ']': '', '`': '', '~': '', '|': '', '\\': '',
        '?': '', '.': '', ',': '', '<': '', '>': '', '«': '', '»': '', '№': '', ':': '', ';': '',
    }

    for k, v in mapping.items():
        s = s.replace(k, v)

    s = transliterate(s.lower())
    s = re.sub('/{2,}', '/', s)
    s = re.sub('[^a-zA-Z0-9/]', '-', s)
    s = re.sub('-{2,}', '-', s)
    s = re.sub('(^-|-$)', '', s)
    s = re.sub('/', '-', s)

    return s


def excel_get_data_names(sheet: Worksheet) -> List[str]:
    r = []

    if sheet.max_row < EXCEL_DATA_NAMES_ROW:
        raise RuntimeWarning("WARNING: sheet '{}' contains too little data".format(sheet.title))

    for col_i in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=EXCEL_DATA_NAMES_ROW, column=col_i).value
        if cell_value:
            r.append(cell_value)
        else:
            break

    return r


def excel_load_sheet(sheet: Worksheet) -> List[Dict]:
    print("Processing sheet '{}'".format(sheet.title))

    if sheet.max_row <= EXCEL_DATA_START_ROW:
        print("WARNING: sheet '{}' contains too little data".format(sheet.title))

    data_names = excel_get_data_names(sheet)

    r = []
    for row_i in range(EXCEL_DATA_START_ROW, sheet.max_row):
        row_data = {}
        for col_i in range(1, len(data_names) + 1):
            data_name_i = col_i - 1
            cell_value = sheet.cell(row=row_i, column=col_i).value

            if isinstance(cell_value, str):
                cell_value = many_spaces_re.sub(' ', cell_value.strip())
            row_data[data_names[data_name_i]] = cell_value

        r.append(row_data)

    print("{} rows successfully loaded from sheet {}".format(len(r), sheet.title))

    return r


def excel_load_workbook(path: str):
    r = []
    wbook = load_workbook(path)

    for sheet in wbook:  # type: Worksheet
        r.append({
            'title': sheet.title,
            'data': excel_load_sheet(sheet)
        })

    return r


def db_connect():
    global db_connection

    print('Trying to connect to MySQL server... ', end='')
    try:
        db_connection = mysql.connector.connect(host=DB_HOST, port=DB_PORT, user=DB_USER, password=DB_PASS,
                                                database=DB_NAME)
        print('OK')
    except mysql.connector.errors.ProgrammingError as e:
        print('FAILED')
        print(e)
        sys.exit(-1)


def db_execute(query: str, params: dict = None) -> MySQLCursor:
    cursor = db_connection.cursor()
    cursor.execute(query, params)

    return cursor


def db_count_rows(query: str, params: dict = None) -> int:
    c = db_execute(query, params)
    count = c.rowcount
    c.fetchall()

    return count


def db_fetch_one(query: str, params: dict = None) -> Optional[Tuple]:
    c = db_execute(query, params)

    r = c.fetchone()
    c.close()

    return r


def db_resolve_category_id(category_name: str) -> int:
    r = db_fetch_one('SELECT category_id FROM category_description WHERE name = "{}"'.format(category_name))

    return r[0] if r else None


def db_resolve_attr_group_id(attr_group_name: str) -> int:
    r = db_fetch_one('SELECT attribute_group_id FROM attribute_group_description '
                     'WHERE name = "{}" AND language_id = {}'.format(attr_group_name, OC_LANGUAGE_ID))

    return r[0] if r else None


def db_resolve_attr_id(attr_group_id: int, attr_name: str):
    # Search for attribute
    r = db_fetch_one('SELECT attribute_id FROM attribute_description '
                     'WHERE name = "{}" AND language_id = {}'.format(attr_name, OC_LANGUAGE_ID))
    if not r:
        raise RuntimeError("Unknown attribute: '{}'".format(attr_name))
    attr_id = r[0]

    # Search for link between attribute and its group
    r = db_fetch_one('SELECT attribute_id FROM attribute '
                     'WHERE attribute_id = {} and attribute_group_id = {}'.format(attr_id, attr_group_id))
    if not r:
        raise RuntimeError("Attribute '{}' does not belong to group id '{}'".format(attr_name, attr_group_id))

    return attr_id


def db_resolve_product_id(sku: str) -> int:
    r = db_fetch_one('SELECT product_id FROM product WHERE sku = "{}" LIMIT 1'.format(sku))

    return r[0] if r else None


def db_resolve_manufacturer_id(manufacturer_name: str) -> int:
    r = db_fetch_one('SELECT manufacturer_id FROM manufacturer WHERE name = "{}" LIMIT 1'.format(manufacturer_name))

    return r[0] if r else None


def db_is_product_exists(sku: str) -> bool:
    return bool(db_resolve_product_id(sku))


def db_product_id_exists(product_id: int) -> bool:
    return bool(db_count_rows('SELECT product_id FROM product WHERE product_id = {}'.format(product_id)))


def db_create_product_attributes(product_id: int, attrs: list):
    if not db_product_id_exists(product_id):
        raise RuntimeError('Product with ID == {} is not found'.format(product_id))

    for attr in attrs:
        db_execute('INSERT INTO product_attribute (product_id, attribute_id, language_id, text) '
                   'VALUES ({}, {}, {}, "{}")'.format(product_id, attr['id'], OC_LANGUAGE_ID,
                                                      db_connection.converter.escape(attr['text'])))


def db_delete_product_attributes(product_id: int) -> int:
    return db_execute('DELETE FROM product_attribute WHERE product_id = {}'.format(product_id)).rowcount


def db_create_product(category_id: int, product_data: dict, attrs: list):
    now = datetime.now().strftime('%Y-%m-%d %H:%M')

    # Resolve manufacturer ID
    manufacturer_id = db_resolve_manufacturer_id(product_data['manufacturer'])
    if not manufacturer_id:
        raise RuntimeError("Manufacturer '{}' is not found".format(product_data['manufacturer']))

    # Insert product
    c = db_execute('INSERT INTO product ' \
                   '(model, sku, quantity, stock_status_id, manufacturer_id, shipping, date_available, price, weight_class_id, ' \
                   'length_class_id, subtract, minimum, sort_order, status, date_added, date_modified)' \
                   'VALUES ("{}", "{}", {}, {}, {}, {}, "{}", {}, {}, {}, {}, {}, {}, {}, "{}", "{}")'.format(
        db_connection.converter.escape(product_data['model']),
        db_connection.converter.escape(product_data['sku']),
        OC_PRODUCT_QUANTITY,
        OC_STOCK_STATUS_ID,
        manufacturer_id,
        OC_PRODUCT_SHIPPING,
        now,
        float(product_data['price']) if product_data['price'] else 0.0,  # Price
        OC_WEIGHT_CLASS_ID,
        OC_LENGTH_CLASS_ID,
        OC_PRODUCT_SUBTRACT,
        OC_PRODUCT_MINIMUM,
        1,  # Sort order
        1,  # Status
        now,  # Date added
        now,  # Date modified
    ))

    product_id = c.lastrowid

    # Link to category
    db_execute('INSERT INTO product_to_category (product_id, category_id) VALUES ({}, {})'.
               format(product_id, category_id))

    # Link to store
    db_execute('INSERT INTO product_to_store (product_id, store_id) VALUES ({}, {})'.
               format(product_id, OC_STORE_ID))

    # Link to layout
    db_execute('INSERT INTO product_to_layout (product_id, store_id, layout_id) VALUES ({}, {}, {})'.
               format(product_id, OC_STORE_ID, OC_LAYOUT_ID))

    # URL alias
    db_execute('INSERT INTO url_alias (query, keyword) VALUES ("product_id={}", "{}")'.
               format(product_id, transform_str(product_data['name'])))

    # Description
    db_execute('INSERT INTO product_description (product_id, language_id, name, description, meta_title)' \
               'VALUES ({}, {}, "{}", "{}", "{}")'.format(
        product_id,
        OC_LANGUAGE_ID,
        db_connection.converter.escape(product_data['name']),
        db_connection.converter.escape(product_data.get('description') or ''),
        db_connection.converter.escape(product_data['name']),
    ))

    # Attributes
    db_create_product_attributes(product_id, attrs)

    debug("New product '{}' successfully created with ID == {}".format(product_data['name'], product_id))


def db_update_product(product_data: dict, attrs: list):
    # Resolve product ID
    product_id = db_resolve_product_id(product_data['sku'])
    if not db_product_id_exists(product_id):
        raise RuntimeError('Product with ID == {} is not found'.format(product_id))

    # Resolve manufacturer ID
    manufacturer_id = db_resolve_manufacturer_id(product_data['manufacturer'])
    if not manufacturer_id:
        raise RuntimeError("Manufacturer '{}' is not found".format(product_data['manufacturer']))

    # Update general info
    db_execute('UPDATE product SET model="{}", manufacturer_id={}, price={}, date_modified="{}" '
               'WHERE product_id={}'.format(
        db_connection.converter.escape(product_data['model']),
        manufacturer_id,
        float(product_data['price']) if product_data['price'] else 0.0,  # Price
        datetime.now().strftime('%Y-%m-%d %H:%M'),
        product_id,
    ))

    # Update attributes
    db_delete_product_attributes(product_id)
    db_create_product_attributes(product_id, attrs)

    # Update title and description
    db_execute('UPDATE product_description SET name="{}", description="{}", meta_title="{}" '
               'WHERE product_id={}'.format(
        db_connection.converter.escape(product_data['name']),
        db_connection.converter.escape(product_data.get('description') or ''),
        db_connection.converter.escape(product_data['name']),
        product_id,
    ))

    debug("Existing product '{}' with ID == {} successfully updated".format(product_data['name'], product_id))


def extract_product_attrs(attr_group_name: str, product_data: Dict[str, str]) -> list:
    attr_group_id = db_resolve_attr_group_id(attr_group_name)

    if not attr_group_id:
        raise RuntimeError("Attribute group '{}' is not found".format(attr_group_name))

    r = []
    keys_to_pop = []
    for k, v in product_data.items():
        if not k.startswith('attr_'):
            continue

        attr_name = k.replace('attr_', '')

        attr_id = db_resolve_attr_id(attr_group_id, attr_name)
        if not attr_id:
            raise RuntimeError("Attribute named '{}' is not found in attribute group '{}'".
                               format(attr_name, attr_group_name))

        r.append({
            'group_name': attr_group_name,
            'group_id': attr_group_id,
            'name': attr_name,
            'id': attr_id,
            'text': v,
        })

        keys_to_pop.append(k)

    for k in keys_to_pop:
        product_data.pop(k)

    return r


def process_product(product_data: dict, attr_group_name: str):
    for req_key in ('sku', 'name', 'manufacturer', 'model', 'category'):
        if req_key not in product_data or not product_data[req_key]:
            raise RuntimeError("'{}' is not in product's data or is is empty".format(req_key))

    category_id = db_resolve_category_id(product_data['category'])
    product_attrs = extract_product_attrs(attr_group_name, product_data)

    if db_is_product_exists(product_data['sku']):
        db_update_product(product_data, product_attrs)
    else:
        db_create_product(category_id, product_data, product_attrs)


def usage() -> str:
    return sys.argv[0] + ' EXCEL_FILE'


def main():
    # Connect to the database
    db_connect()

    # Load data from Excel's workbook
    for sheet in excel_load_workbook(sys.argv[1]):
        sheet_title = sheet['title']
        row_i = 0
        for product_data in sheet['data']:
            try:
                process_product(product_data, sheet_title)
                row_i += 1
            except RuntimeError as e:
                raise RuntimeError("Sheet '{}', row {}: {}".format(sheet_title, row_i + EXCEL_DATA_START_ROW, e))


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(usage())
        sys.exit(-1)

    main()
